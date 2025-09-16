# ESUSSifilis.py
# -*- coding: utf-8 -*-
"""
App Streamlit:
- Login (e-mail/senha)
- Carrega até 64 planilhas (CSV/XLS/XLSX/ODS)
- Detecta a UNIDADE (B10 > A3 > varredura 10 linhas)
- Mantém TODAS as colunas
- Filtro por UNIDADE (coluna BA/UNIDADE) com botão
- Relatórios por AR/AS/AT/AU/AV/AN/AJ (mantendo todas as colunas)
- Exporta XLSX com cabeçalho na linha 25:
    * A25 = UNIDADE
    * BA25 = UNIDADE
    * BA (todas as linhas de dados) = UNIDADE quando A (paciente) estiver preenchida
"""

import io
import re
import csv
import unicodedata
import warnings
from io import StringIO
from typing import List, Tuple, Dict, Optional

import pandas as pd
import streamlit as st
from pathlib import Path

# =========================
# CONFIGURAÇÕES
# =========================
st.set_page_config(page_title="e-SUS — Filtros", layout="wide")
warnings.simplefilter("ignore", category=UserWarning)

# =========================
# BANNER NO LOGIN (apenas no login)
# =========================
HERO_CANDIDATES = [
    r"C:\Users\raque\Desktop\Banco Sifilis\assets\gestaweb_ds7.jpg",  # seu caminho
    "assets/gestaweb_ds7.jpg",                                        # relativo ao projeto
    "/mnt/data/gestaweb_ds7.jpg",                                     # fallback
]

def _pick_hero_path() -> Optional[Path]:
    for p in HERO_CANDIDATES:
        path = Path(p)
        if path.exists():
            return path
    return None

def show_login_banner():
    hero = _pick_hero_path()
    if hero and hero.exists():
        st.image(str(hero), use_container_width=True)
    else:
        st.info(
            "Imagem do banner não encontrada. Verifique o caminho em HERO_CANDIDATES "
            "ou copie o arquivo para ./assets/gestaweb_ds7.jpg"
        )

# =========================
# NOMES DE COLUNAS E FILTROS
# =========================
COLMAP = {
    "paciente": [
        "nome do paciente", "paciente", "nome",
        "nome do usuário", "nome do usuario",
        "nome da usuaria", "nome da usuária",
        "usuario", "usuário"
    ],
}

COL_SPECS = {
    "AR": {"desc": "Exame de HIV no 1º trimestre", "type": "text_nao",
           "names": ["Exame de HIV no primeiro trimestre", "hiv 1º trimestre", "hiv 1 tri"]},
    "AS": {"desc": "Exame de Sífilis no 1º trimestre", "type": "text_nao",
           "names": ["Exame de Sífilis no primeiro trimestre", "sífilis 1º trimestre", "sifilis 1 tri"]},
    "AT": {"desc": "Exame de Hepatite B no 1º trimestre", "type": "text_nao",
           "names": ["Exame de Hepatite B no primeiro trimestre", "hbv 1º trimestre", "hepatite b 1 tri"]},
    "AU": {"desc": "Exame de Hepatite C no 1º trimestre", "type": "text_nao",
           "names": ["Exame de Hepatite C no primeiro trimestre", "hcv 1º trimestre", "hepatite c 1 tri"]},
    "AV": {"desc": "Exame de HIV no 3º trimestre", "type": "text_nao",
           "names": ["Exame de HIV no terceiro trimestre", "hiv 3º trimestre", "hiv 3 tri"]},
    "AN": {"desc": "Qtde atendimentos odontológicos no pré-natal", "type": "num_lt", "value": 1,
           "names": ["Quantidade de atendimentos odontológicos no pré-natal", "odontol", "atend odont"]},
    "AJ": {"desc": "IG (DUM) (semanas)", "type": "num_gt", "value": 43,
           "names": ["IG (DUM) (semanas)", "idade gestacional (dum)", "ig semanas"]},
}

# =========================
# AUTENTICAÇÃO
# =========================
def get_allowed_credentials() -> Tuple[str, str]:
    try:
        email = st.secrets["auth"]["email"]
        pwd = st.secrets["auth"]["password"]
    except Exception:
        email = "vigilanciaepidemiologicadsvii@gmail.com"
        pwd = "epidemiosifilis"
    return email, pwd

def login_block():
    # banner só no login
    show_login_banner()

    st.title("🔐 Login — e-SUS")
    with st.form("login_form", clear_on_submit=False):
        email_in = st.text_input("E-mail", value="", placeholder="seu@email")
        pwd_in = st.text_input("Senha", value="", type="password")
        ok = st.form_submit_button("Entrar")

    if ok:
        allowed_email, allowed_pwd = get_allowed_credentials()
        if email_in.strip().lower() == allowed_email.lower() and pwd_in == allowed_pwd:
            st.session_state["auth_ok"] = True
            st.session_state["user_email"] = email_in.strip()
            try:
                st.rerun()
            except Exception:
                st.experimental_rerun()
        else:
            st.error("E-mail ou senha inválidos.")
    st.stop()

# =========================
# FUNÇÕES AUXILIARES (ETL)
# =========================
def normalize(s: str) -> str:
    if not isinstance(s, str):
        s = str(s)
    s = s.strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s

def find_first_matching_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols_norm = {normalize(c): c for c in df.columns}
    for c in candidates:
        nc = normalize(c)
        if nc in cols_norm:
            return cols_norm[nc]
    for real in df.columns:
        rn = normalize(real)
        for cand in candidates:
            if normalize(cand) in rn:
                return real
    return None

def _read_csv_robusto(file_obj) -> pd.DataFrame:
    if hasattr(file_obj, "read"):
        pos = file_obj.tell() if hasattr(file_obj, "tell") else None
        data = file_obj.read()
        if pos is not None:
            try:
                file_obj.seek(pos)
            except Exception:
                pass
    else:
        with open(file_obj, "rb") as fh:
            data = fh.read()

    if isinstance(data, bytes):
        try:
            text = data.decode("utf-8", "ignore")
        except Exception:
            text = data.decode("latin-1", "ignore")
    else:
        text = str(data)

    buf = StringIO(text)
    try:
        return pd.read_csv(buf, sep=None, engine="python", header=None, dtype=str,
                           skip_blank_lines=True, quoting=csv.QUOTE_MINIMAL, on_bad_lines="skip")
    except Exception:
        pass

    for sep in [";", "\t", ",", "|"]:
        buf.seek(0)
        try:
            return pd.read_csv(buf, sep=sep, engine="python", header=None, dtype=str,
                               skip_blank_lines=True, quoting=csv.QUOTE_MINIMAL, on_bad_lines="skip")
        except Exception:
            continue

    buf.seek(0)
    return pd.read_csv(buf, header=None, dtype=str, engine="python", on_bad_lines="skip")

def detect_unit_name(raw_df: pd.DataFrame) -> Optional[str]:
    try:
        b10 = raw_df.iloc[9, 1]
        if isinstance(b10, str) and b10.strip() != "":
            return b10.strip()
    except Exception:
        pass

    try:
        a3 = raw_df.iloc[2, 0]
        if isinstance(a3, str) and normalize(a3).startswith("unidade de saude"):
            return a3.strip()
    except Exception:
        pass

    for i in range(min(10, len(raw_df))):
        try:
            v = str(raw_df.iloc[i, 0])
            if re.search(r"(?i)unidade\s+de\s+saud(e|é)", v):
                return v.strip()
        except Exception:
            continue
    return None

def read_any_table(file) -> Tuple[pd.DataFrame, Optional[str]]:
    name = getattr(file, "name", str(file)).lower()

    if name.endswith(".csv"):
        raw = _read_csv_robusto(file)
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        raw = pd.read_excel(file, header=None, dtype=str)
    elif name.endswith(".ods"):
        try:
            raw = pd.read_excel(file, engine="odf", header=None, dtype=str)
        except Exception:
            raw = pd.read_excel(file, header=None, dtype=str)
    else:
        raw = _read_csv_robusto(file)

    unidade = detect_unit_name(raw)

    # Detecta linha de cabeçalho
    header_row = None
    for i in range(min(30, len(raw))):
        row = raw.iloc[i]
        non_na = row.dropna().astype(str)
        text_count = sum(len(str(x).strip()) > 0 for x in non_na)
        if text_count >= 4:
            header_row = i
            break
    if header_row is None:
        header_row = 0

    # Reconstrói DF
    df = raw.copy()
    df.columns = raw.iloc[header_row].fillna("")
    df = raw.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = [str(c).strip() if str(c).strip() != "" else f"col_{i}" for i, c in enumerate(df.columns)]

    # Mantém linhas com paciente
    col_paciente = find_first_matching_col(df, COLMAP["paciente"]) or df.columns[0]
    df = df[df[col_paciente].notna() & (df[col_paciente].astype(str).str.strip() != "")]
    df["UNIDADE"] = unidade if unidade else "(UNIDADE NAO DETECTADA)"

    return df, unidade

def excel_letter_to_index(letter: str) -> int:
    if not letter or not letter.isalpha():
        return -1
    letter = letter.upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1

def get_col_by_letter_or_name(df: pd.DataFrame, letter: str, names: List[str]) -> Optional[str]:
    j = excel_letter_to_index(letter)
    if 0 <= j < len(df.columns):
        col = df.columns[j]
        if str(col).strip() != "":
            return col
    names_n = [normalize(n) for n in names]
    for c in df.columns:
        cn = normalize(c)
        for nn in names_n:
            if cn == nn or nn in cn:
                return c
    return None

def series_is_nao(s: pd.Series) -> pd.Series:
    sn = s.astype(str).map(normalize)
    return sn.str.fullmatch(r"n[aã]o")

def to_numeric_series(s: pd.Series) -> pd.Series:
    ss = s.astype(str).str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
    return pd.to_numeric(ss, errors="coerce")

def build_requested_filters(df: pd.DataFrame) -> Tuple[Dict[str, pd.Series], Dict[str, str]]:
    masks, found_cols = {}, {}
    for code, spec in COL_SPECS.items():
        col = get_col_by_letter_or_name(df, code, spec.get("names", []))
        if col is None:
            continue
        found_cols[code] = col
        t = spec["type"]
        if t == "text_nao":
            masks[code] = series_is_nao(df[col])
        elif t == "num_lt":
            v = spec.get("value", 0)
            x = to_numeric_series(df[col])
            masks[code] = x < v
        elif t == "num_gt":
            v = spec.get("value", 0)
            x = to_numeric_series(df[col])
            masks[code] = x > v
    return masks, found_cols

def write_xlsx_with_ba_layout(sheets: Dict[str, pd.DataFrame]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    HEADER_ROW = 25
    DATA_START_ROW = 26
    COL_A_IDX = 1
    COL_BA_IDX = 53

    for sheet_name, df_in in sheets.items():
        col_paciente = find_first_matching_col(df_in, COLMAP["paciente"]) or df_in.columns[0]
        cols_order = [col_paciente] + [c for c in df_in.columns if c != col_paciente]
        df = df_in[cols_order].copy()

        ws = wb.create_sheet(title=sheet_name[:31])

        for j, col in enumerate(df.columns, start=1):
            ws.cell(row=HEADER_ROW, column=j, value=str(col))

        unidade_val = "(SEM COLUNA UNIDADE)"
        if "UNIDADE" in df.columns:
            uniques = sorted(set(df["UNIDADE"].dropna().astype(str)))
            unidade_val = uniques[0] if len(uniques) == 1 else "(VARIAS UNIDADES)"
        ws.cell(row=HEADER_ROW, column=COL_A_IDX, value=unidade_val)
        ws.cell(row=HEADER_ROW, column=COL_BA_IDX, value=unidade_val)

        for i, (_, row) in enumerate(df.iterrows(), start=DATA_START_ROW):
            for j, col in enumerate(df.columns, start=1):
                val = None if pd.isna(row[col]) else str(row[col])
                ws.cell(row=i, column=j, value=val)

            val_a = ws.cell(row=i, column=COL_A_IDX).value
            if val_a is not None and str(val_a).strip() != "":
                ws.cell(row=i, column=COL_BA_IDX, value=unidade_val)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")

def single_sheet_xlsx(df: pd.DataFrame, sheet_name: str = "PLANILHA") -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    out.seek(0)
    return out.read()

# =========================
# UI STREAMLIT
# =========================
# Gate de login
if not st.session_state.get("auth_ok"):
    login_block()

# Sidebar (logout)
with st.sidebar:
    if st.session_state.get("auth_ok"):
        st.success(f"Conectado: {st.session_state.get('user_email','')}")
        if st.button("Sair"):
            st.session_state.clear()
            try:
                st.rerun()
            except Exception:
                st.experimental_rerun()

st.title("Unidades de Saúde DSVII — e-SUS Monitoramento")

# Upload de arquivos
uploaded_files = st.file_uploader(
    "Selecione as planilhas (até 64)",
    type=["csv", "xlsx", "xls", "ods"],
    accept_multiple_files=True,
)

if uploaded_files:
    if len(uploaded_files) > 64:
        st.warning("Você selecionou mais de 64 arquivos. Apenas os 64 primeiros serão processados.")
        uploaded_files = uploaded_files[:64]

    with st.spinner("Lendo e preparando as planilhas..."):
        dfs, unidades_detectadas = [], []
        for f in uploaded_files:
            try:
                df, unidade = read_any_table(f)
                dfs.append(df)
                unidades_detectadas.append(unidade or "(não detectada)")
            except Exception as e:
                st.error(f"Erro ao ler {getattr(f, 'name', 'arquivo')}: {e}")
        if not dfs:
            st.stop()
        base = pd.concat(dfs, ignore_index=True)

    # ---- Visualização geral ----
    st.subheader("Pré-visualização da base (todas as colunas)")
    st.dataframe(base.head(300), use_container_width=True, height=420)

    # ---- Filtro por UNIDADE (coluna BA/UNIDADE) com botão ----
    st.subheader("Filtro por UNIDADE (coluna BA/UNIDADE)")
    units = sorted(base["UNIDADE"].dropna().astype(str).unique()) if "UNIDADE" in base.columns else []
    sel = st.multiselect("Selecione a(s) UNIDADE(s)", options=units, default=units[:1] if units else [])

    if "unit_filter_on" not in st.session_state:
        st.session_state["unit_filter_on"] = False

    c1, c2, _ = st.columns([1, 1, 3])
    with c1:
        if st.button("Aplicar filtro por UNIDADE (BA)"):
            st.session_state["unit_filter_on"] = True
    with c2:
        if st.button("Limpar filtro"):
            st.session_state["unit_filter_on"] = False

    base_filtrada = base.copy()
    if st.session_state["unit_filter_on"] and sel:
        base_filtrada = base_filtrada[base_filtrada["UNIDADE"].isin(sel)]

    st.write(f"Linhas após filtro por UNIDADE: {len(base_filtrada)}")
    st.dataframe(base_filtrada.head(300), use_container_width=True, height=420)

    # >>> Fonte única para relatórios (respeita o filtro quando ligado)
    df_fonte_relatorios = base_filtrada if (st.session_state["unit_filter_on"] and sel) else base

    # ---- Relatórios por AR/AS/AT/AU/AV/NA/AJ ----
    st.subheader("Relatórios específicos (mantendo TODAS as colunas)")
    masks, found_cols = build_requested_filters(df_fonte_relatorios)

    if found_cols:
        st.caption("Colunas detectadas (código → nome real): " +
                   ", ".join([f"{k}→{v}" for k, v in found_cols.items()]))
    else:
        st.info("AR/AS/AT/AU/AV/AN/AJ não localizadas por letra nem por nome — verifique o cabeçalho.")

    labels = {
        "AR": "HIV 1º tri = NÃO",
        "AS": "Sífilis 1º tri = NÃO",
        "AT": "Hepatite B 1º tri = NÃO",
        "AU": "Hepatite C 1º tri = NÃO",
        "AV": "HIV 3º tri = NÃO",
        "AN": "Odonto pré-natal < 1",
        "AJ": "IG (DUM) semanas > 43",
    }

    tables_spec: Dict[str, pd.DataFrame] = {}
    for code in ["AR", "AS", "AT", "AU", "AV", "AN", "AJ"]:
        if code in masks:
            df_tab = df_fonte_relatorios[masks[code]].copy()
            tables_spec[code] = df_tab
            with st.expander(f"{labels[code]} — {len(df_tab)} linha(s)"):
                st.dataframe(df_tab.head(300), use_container_width=True, height=360)
                cdl1, cdl2 = st.columns(2)
                with cdl1:
                    st.download_button(
                        f"Baixar {code}.csv",
                        data=df_to_csv_bytes(df_tab),
                        file_name=f"{code}.csv",
                        mime="text/csv",
                        key=f"csv_{code}"
                    )
                with cdl2:
                    st.download_button(
                        f"Baixar {code}.xlsx",
                        data=single_sheet_xlsx(df_tab, sheet_name=code),
                        file_name=f"{code}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"xlsx_{code}"
                    )
        else:
            st.info(f"Filtro {labels.get(code, code)}: coluna {code} não localizada.")

      

   
else:
    st.info("Carregue seus arquivos para iniciar. Formatos aceitos: CSV, XLSX/XLS e ODS (requer odfpy).")
